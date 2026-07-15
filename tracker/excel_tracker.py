from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

COLUMNS = [
    "Company",
    "Role",
    "Platform",
    "Location",
    "Job Type",
    "Salary",
    "Match %",
    "Resume Used",
    "Applied",
    "Interview",
    "OA",
    "HR",
    "Offer",
    "Rejected",
    "Notes",
    "Job URL",
]
_URL_COL = COLUMNS.index("Job URL")


def ensure_workbook(path: str) -> None:
    if not Path(path).exists():
        Path(path).parent.mkdir(parents=True, exist_ok=True)
        wb = Workbook()
        ws = wb.active
        ws.title = "Applications"
        ws.append(COLUMNS)
        wb.save(path)
        return

    _migrate_schema(path)


def _migrate_schema(path: str) -> None:
    """Rebuild the sheet under the current COLUMNS if the on-disk header is stale.

    Preserves every existing row (including any manual notes/statuses you've
    typed in) by column name, so adding a new column later doesn't require
    deleting your tracker -- new columns just come out blank for old rows.
    """
    wb = load_workbook(path)
    ws: Worksheet = wb["Applications"]
    header = [cell.value for cell in ws[1]]
    if header == COLUMNS:
        return

    old_rows = [
        {header[i]: cell.value for i, cell in enumerate(row) if i < len(header)}
        for row in ws.iter_rows(min_row=2)
    ]

    ws.delete_rows(1, ws.max_row)
    ws.append(COLUMNS)
    for old_row in old_rows:
        ws.append([old_row.get(col, "") for col in COLUMNS])
    wb.save(path)


def upsert_row(path: str, row: dict) -> None:
    """Insert a row for this job's URL, or update the existing one in place if already synced.

    Keyed on URL rather than (Company, Role): distinct postings can share an
    identical title at the same company (e.g. multiple "Accenture - Custom
    Software Engineer" listings from one Naukri search), which used to
    silently collapse into a single row and lose real job records.
    """
    ensure_workbook(path)
    wb = load_workbook(path)
    ws: Worksheet = wb["Applications"]

    for existing in ws.iter_rows(min_row=2):
        if existing[_URL_COL].value == row.get("Job URL"):
            for i, col in enumerate(COLUMNS):
                if col in row:
                    existing[i].value = row[col]
            wb.save(path)
            return

    ws.append([row.get(col, "") for col in COLUMNS])
    wb.save(path)


def update_status(path: str, job_url: str, column: str, value: str) -> bool:
    """Find the row matching this job's URL and update one status column. Returns True if found."""
    wb = load_workbook(path)
    ws: Worksheet = wb["Applications"]
    col_index = COLUMNS.index(column)

    for row in ws.iter_rows(min_row=2):
        if row[_URL_COL].value == job_url:
            row[col_index].value = value
            wb.save(path)
            return True
    return False
